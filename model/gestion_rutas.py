import os, io, csv, json
import xml.etree.ElementTree as ET
from psycopg2.extras import RealDictCursor
from psycopg2 import errors
from datetime import datetime
from dotenv import load_dotenv
from zoneinfo import ZoneInfo
# Importar la función para obtener el pool de conexiones
from database.database_manager import _get_pool as get_db_pool

try:
    import openpyxl
except Exception:
    openpyxl = None

load_dotenv()
TZ_BOGOTA = ZoneInfo("America/Bogota")
KML_NS = "http://www.opengis.net/kml/2.2"

def ahora_bogota() -> datetime:
    return datetime.now(TZ_BOGOTA)

# ═══════════════════════════════════════════════════════════════════════════════
# Helpers KML — parseo puro, sin dependencia de BD ni PostGIS
# ═══════════════════════════════════════════════════════════════════════════════
def _parse_coord_string(texto: str) -> list:
    """Convierte 'lon,lat,alt lon,lat,alt ...' a lista de [lon, lat]."""
    puntos = []
    for token in texto.strip().split():
        partes = token.split(",")
        if len(partes) >= 2:
            try:
                puntos.append([float(partes[0]), float(partes[1])])
            except ValueError:
                pass
    return puntos

def _inferir_variante(nombre: str) -> int:
    up = (nombre or "").upper()
    if "V2" in up or "VUELTA" in up or "_SN_" in up:
        return 2
    if "V3" in up:
        return 3
    if "V4" in up:
        return 4
    return 1

def _extraer_codigo_y_nombre(nombre_placemark: str):
    """De '1. 704A01_Altos de Serrezuela' extrae ('704A01', 'Altos de Serrezuela')."""
    partes = nombre_placemark.split(".")
    resto = partes[-1].strip() if len(partes) > 1 else nombre_placemark
    if "_" in resto:
        idx = resto.index("_")
        return resto[:idx].strip(), resto[idx + 1:].strip()
    return None, resto.strip()

def parsear_kml_bytes(contenido_bytes: bytes) -> list:
    """
    Parsea un KML y retorna lista de dicts con:
        - nombre_placemark, tipo ('trazado'|'paradero')
        - coordenadas_json: GeoJSON compacto con todos los metadatos embebidos
        - total_puntos

    Esquema unificado de coordenadas_json:
        Trazado:  {"type":"LineString","coordinates":[[lon,lat],...],"variante":1}
        Paradero: {"type":"Point","coordinates":[lon,lat],
                    "secuencia":1,"codigo":"704A01","nombre":"Altos de Serrezuela","variante":1}

    Sin columnas separadas lat/lon/secuencia/codigo_parada/nombre_parada/num_variante.
    """
    root = ET.fromstring(contenido_bytes)
    ns = {"k": KML_NS}
    placemarks = root.findall(".//k:Placemark", ns)
    registros = []
    seq_por_variante = {}

    for pm in placemarks:
        nombre_el = pm.find("k:name", ns)
        nombre = nombre_el.text.strip() if (nombre_el is not None and nombre_el.text) else ""

        # ── Trazado: LineString ──────────────────────────────────────────────
        coords_el = pm.find(".//k:LineString/k:coordinates", ns)
        if coords_el is not None and coords_el.text:
            puntos = _parse_coord_string(coords_el.text)
            if len(puntos) < 2:
                continue
            variante = _inferir_variante(nombre)
            registros.append({
                "nombre_placemark": nombre,
                "tipo": "trazado",
                "coordenadas_json": json.dumps({
                    "type": "LineString",
                    "coordinates": puntos,
                    "variante": variante,
                }, separators=(',', ':')),
                "total_puntos": len(puntos),
            })
            continue

        # ── Paradero: Point ──────────────────────────────────────────────────
        coords_el = pm.find(".//k:Point/k:coordinates", ns)
        if coords_el is not None and coords_el.text:
            puntos = _parse_coord_string(coords_el.text)
            if not puntos:
                continue
            lon, lat = puntos[0]
            variante = _inferir_variante(nombre)
            seq_por_variante.setdefault(variante, 0)
            seq_por_variante[variante] += 1
            codigo, nombre_parada = _extraer_codigo_y_nombre(nombre)

            seq_num = seq_por_variante[variante]
            partes_nombre = nombre.split(".")
            if len(partes_nombre) > 1:
                try:
                    seq_num = int(partes_nombre[0].strip())
                except ValueError:
                    pass

            registros.append({
                "nombre_placemark": nombre,
                "tipo": "paradero",
                "coordenadas_json": json.dumps({
                    "type": "Point",
                    "coordinates": [lon, lat],
                    "secuencia": seq_num,
                    "codigo": codigo,
                    "nombre": nombre_parada,
                    "variante": variante,
                }, separators=(',', ':')),
                "total_puntos": 1,
            })

    return registros

# ═══════════════════════════════════════════════════════════════════════════════
# DDL config.rutas_kml — SIN PostGIS, 100 % PostgreSQL nativo
# ═══════════════════════════════════════════════════════════════════════════════
_DDL_RUTAS_KML = """
CREATE TABLE IF NOT EXISTS config.rutas_kml (
    id               BIGSERIAL    PRIMARY KEY,
    id_linea         BIGINT       NOT NULL,
    nombre_placemark TEXT,
    tipo             VARCHAR(10)  NOT NULL CHECK (tipo IN ('trazado','paradero')),
    coordenadas_json TEXT         NOT NULL,
    total_puntos     INT,
    archivo_origen   TEXT,
    cargado_por      TEXT,
    created_at       TIMESTAMPTZ  NOT NULL DEFAULT now(),
    updated_at       TIMESTAMPTZ  NOT NULL DEFAULT now()
);
CREATE INDEX IF NOT EXISTS idx_rutas_kml_id_linea
    ON config.rutas_kml (id_linea);
CREATE INDEX IF NOT EXISTS idx_rutas_kml_linea_tipo
    ON config.rutas_kml (id_linea, tipo);
"""

# ═══════════════════════════════════════════════════════════════════════════════
# GestionRutas — CRUD de config.rutas
# ═══════════════════════════════════════════════════════════════════════════════
class GestionRutas:
    """
    Gestión CRUD de la tabla config.rutas.
    Estructura esperada:
        config.rutas (
            id           BIGSERIAL PRIMARY KEY,
            id_linea     BIGINT NOT NULL,
            ruta_comercial TEXT NOT NULL,
            id_cop       BIGINT NOT NULL REFERENCES config.cop(id),
            estado       SMALLINT NOT NULL DEFAULT 1,
            created_at   TIMESTAMPTZ DEFAULT now(),
            updated_at   TIMESTAMPTZ DEFAULT now()
        )
    """

    def __init__(self):
        self.connection = get_db_pool().getconn()
        if not self.connection.closed:
            self.connection.rollback()
        self.connection.cursor_factory = RealDictCursor
        self.cursor = self.connection.cursor()

    def cerrar_conexion(self):
        if getattr(self, "cursor", None):
            try:
                self.cursor.close()
            except Exception:
                pass
            self.cursor = None
        if getattr(self, "connection", None):
            try:
                if not self.connection.closed:
                    self.connection.rollback()
                get_db_pool().putconn(self.connection)
            except Exception:
                pass
            self.connection = None

    # ── Helpers ─────────────────────────────────────────────────────────────
    def _col_exists(self, schema: str, table: str, column: str) -> bool:
        self.cursor.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema=%s AND table_name=%s AND column_name=%s
            LIMIT 1
            """,
            (schema, table, column),
        )
        return self.cursor.fetchone() is not None

    def _cop_joins_and_exprs(self):
        """Devuelve (joins_sql, comp_expr, zona_expr) según la estructura real de config.cop."""
        has_id_comp = self._col_exists("config", "cop", "id_componente")
        has_id_zona = self._col_exists("config", "cop", "id_zona")
        joins, comp_expr, zona_expr = [], "NULL", "NULL"
        if has_id_comp:
            joins.append("LEFT JOIN config.componente comp ON comp.id = c.id_componente")
            comp_expr = "comp.componente"
        if has_id_zona:
            joins.append("LEFT JOIN config.zona z ON z.id = c.id_zona")
            zona_expr = "z.zona"
        return (" ".join(joins), comp_expr, zona_expr)

    def _fila_o_error(self, cur):
        fila = cur.fetchone()
        if not fila:
            raise ValueError("Registro no encontrado")
        return fila

    def _paginacion(self, pagina: int, tamano: int):
        off = (pagina - 1) * tamano
        return " LIMIT %s OFFSET %s ", [tamano, off]

    def _filtro_busqueda(self, q: str):
        if not q:
            return "", []
        qn = f"%{q.strip().upper()}%"
        return (
            " AND (UPPER(r.ruta_comercial) LIKE %s OR CAST(r.id_linea AS TEXT) LIKE %s) ",
            [qn, qn],
        )

    def _u(self, s):
        if s is None:
            return None
        return str(s).strip().upper()

    def _resolver_id_cop(self, id_cop=None, nombre_cop=None):
        if id_cop:
            return int(id_cop)
        if nombre_cop:
            self.cursor.execute(
                "SELECT id FROM config.cop WHERE UPPER(cop) = %s",
                (nombre_cop.strip().upper(),),
            )
            row = self.cursor.fetchone()
            if not row:
                raise ValueError(f"COP '{nombre_cop}' no existe")
            return int(row["id"])
        raise ValueError("Centro de Operación no especificado (ID COP o nombre)")

    # ── Listar rutas ─────────────────────────────────────────────────────────
    def listar_rutas(
        self,
        q=None,
        estado=None,
        id_cop=None,
        componente=None,
        zona=None,
        pagina: int = 1,
        tamano: int = 10,
    ):
        where, params = " WHERE 1=1 ", []

        s, sp = self._filtro_busqueda(q)
        where += s
        params += sp

        if estado is not None:
            where += " AND r.estado=%s "
            params.append(int(estado))
        if id_cop:
            where += " AND r.id_cop=%s "
            params.append(int(id_cop))

        joins_cop, comp_expr, zona_expr = self._cop_joins_and_exprs()

        if componente:
            where += f" AND UPPER({comp_expr})=%s "
            params.append(self._u(componente))
        if zona:
            where += f" AND UPPER({zona_expr})=%s "
            params.append(self._u(zona))

        sql_count = f"""
            SELECT COUNT(1)
            FROM config.rutas r
            JOIN config.cop c ON c.id = r.id_cop
            {joins_cop}
            {where}
        """
        self.cursor.execute(sql_count, params)
        total = self.cursor.fetchone()["count"]

        order = " ORDER BY r.id_linea ASC, r.ruta_comercial ASC "
        pag, pp = self._paginacion(pagina, tamano)
        sqlq = f"""
            SELECT
                r.id,
                r.id_linea,
                r.ruta_comercial,
                r.estado,
                r.id_cop,
                c.cop,
                {comp_expr} AS componente,
                {zona_expr} AS zona,
                to_char(r.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                to_char(r.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM config.rutas r
            JOIN config.cop c ON c.id = r.id_cop
            {joins_cop}
            {where} {order} {pag}
        """
        self.cursor.execute(sqlq, params + pp)
        return self.cursor.fetchall(), total

    # ── CRUD ─────────────────────────────────────────────────────────────────
    def crear_ruta(
        self,
        id_linea: int,
        ruta_comercial: str,
        id_cop: int,
        estado: int = 1,
    ):
        ruta_u = self._u(ruta_comercial)
        if not ruta_u:
            raise ValueError("La ruta comercial es obligatoria")
        if not id_linea:
            raise ValueError("El ID de línea es obligatorio")

        try:
            self.cursor.execute(
                """
                INSERT INTO config.rutas (id_linea, ruta_comercial, id_cop, estado)
                VALUES (%s, %s, %s, %s)
                RETURNING id, id_linea, ruta_comercial, estado, id_cop,
                        to_char(created_at,'YYYY-MM-DD HH24:MI') AS created_at,
                        to_char(updated_at,'YYYY-MM-DD HH24:MI') AS updated_at
                """,
                (int(id_linea), ruta_u, int(id_cop), int(estado)),
            )
            fila = self.cursor.fetchone()
            self.connection.commit()
            return fila
        except errors.ForeignKeyViolation:
            self.connection.rollback()
            raise ValueError("El Centro de Operación (COP) no existe")
        except errors.UniqueViolation:
            self.connection.rollback()
            raise ValueError("Ya existe una ruta con ese ID de línea y ruta comercial")
        except Exception as e:
            self.connection.rollback()
            raise e

    def actualizar_ruta(
        self,
        id: int,
        id_linea: int,
        ruta_comercial: str,
        id_cop: int,
        estado: int = 1,
    ):
        ruta_u = self._u(ruta_comercial)
        if not ruta_u:
            raise ValueError("La ruta comercial es obligatoria")
        if not id_linea:
            raise ValueError("El ID de línea es obligatorio")

        try:
            self.cursor.execute(
                """
                UPDATE config.rutas
                SET id_linea=%s, ruta_comercial=%s, id_cop=%s, estado=%s, updated_at=now()
                WHERE id=%s
                RETURNING id, id_linea, ruta_comercial, estado, id_cop,
                        to_char(created_at,'YYYY-MM-DD HH24:MI') AS created_at,
                        to_char(updated_at,'YYYY-MM-DD HH24:MI') AS updated_at
                """,
                (int(id_linea), ruta_u, int(id_cop), int(estado), int(id)),
            )
            fila = self._fila_o_error(self.cursor)
            self.connection.commit()
            return fila
        except errors.ForeignKeyViolation:
            self.connection.rollback()
            raise ValueError("El Centro de Operación (COP) no existe")
        except errors.UniqueViolation:
            self.connection.rollback()
            raise ValueError("Ya existe una ruta con ese ID de línea y ruta comercial")
        except Exception as e:
            self.connection.rollback()
            raise e

    def cambiar_estado(self, id: int, estado: int):
        self.cursor.execute(
            """
            UPDATE config.rutas
            SET estado=%s, updated_at=now()
            WHERE id=%s
            RETURNING id, id_linea, ruta_comercial, estado,
                    to_char(created_at,'YYYY-MM-DD HH24:MI') AS created_at,
                    to_char(updated_at,'YYYY-MM-DD HH24:MI') AS updated_at
            """,
            (int(estado), int(id)),
        )
        fila = self._fila_o_error(self.cursor)
        self.connection.commit()
        return fila

    # ── Datos de apoyo ────────────────────────────────────────────────────────
    def listar_cop(self, componente: str = None, zona: str = None, estado: int = 1):
        joins_cop, comp_expr, zona_expr = self._cop_joins_and_exprs()
        where, params = " WHERE 1=1 ", []
        if estado is not None:
            where += " AND c.estado=%s "
            params.append(int(estado))
        if componente:
            where += f" AND UPPER({comp_expr})=%s "
            params.append(self._u(componente))
        if zona:
            where += f" AND UPPER({zona_expr})=%s "
            params.append(self._u(zona))
        sql = f"""
            SELECT c.id, c.cop,
                {comp_expr} AS componente,
                {zona_expr} AS zona
            FROM config.cop c
            {joins_cop}
            {where}
            ORDER BY {comp_expr}, {zona_expr}, c.cop
        """
        self.cursor.execute(sql, params)
        return self.cursor.fetchall()

    def listar_componentes(self):
        joins_cop, comp_expr, _ = self._cop_joins_and_exprs()
        sql = f"""
            SELECT DISTINCT {comp_expr} AS componente
            FROM config.cop c
            {joins_cop}
            WHERE c.estado=1 AND {comp_expr} IS NOT NULL
            ORDER BY {comp_expr}
        """
        self.cursor.execute(sql)
        return [r["componente"] for r in self.cursor.fetchall() if r["componente"]]

    def listar_zonas(self, componente: str = None):
        joins_cop, comp_expr, zona_expr = self._cop_joins_and_exprs()
        where, params = " WHERE c.estado=1 ", []
        if componente:
            where += f" AND UPPER({comp_expr})=%s "
            params.append(self._u(componente))
        sql = f"""
            SELECT DISTINCT {zona_expr} AS zona
            FROM config.cop c
            {joins_cop}
            {where} AND {zona_expr} IS NOT NULL
            ORDER BY {zona_expr}
        """
        self.cursor.execute(sql, params)
        return [r["zona"] for r in self.cursor.fetchall() if r["zona"]]

    # ── Normalización para carga masiva ───────────────────────────────────────
    def _normalizar_payload_archivo(self, d: dict):
        def get(k):
            for key, val in d.items():
                norm = key.strip().replace(" ", "_").upper() if key else ""
                if norm == k.strip().replace(" ", "_").upper():
                    return None if val is None else str(val).strip()
            return None

        id_linea_raw = get("Id_linea") or get("ID_Linea") or get("idlinea")
        if id_linea_raw:
            if id_linea_raw.endswith(".0"):
                id_linea_raw = id_linea_raw[:-2]
        if not id_linea_raw or not id_linea_raw.isdigit():
            raise ValueError(f"Id_linea inválido: '{id_linea_raw}'")

        ruta_raw = get("Ruta_comercial") or get("Ruta_Comercial") or get("Ruta")
        if not ruta_raw:
            raise ValueError("Ruta_comercial vacía")

        cop_raw = get("cop") or get("Centro_Operacion") or get("ID_COP") or get("id_cop")
        if not cop_raw:
            raise ValueError("Campo COP / Centro Operacion vacío")
        if cop_raw.replace(".", "").isdigit():
            val = cop_raw
            if val.endswith(".0"):
                val = val[:-2]
            id_cop = int(val)
        else:
            id_cop = self._resolver_id_cop(nombre_cop=cop_raw)

        return {
            "id_linea": int(id_linea_raw),
            "ruta_comercial": self._u(ruta_raw),
            "id_cop": id_cop,
            "estado": 1,
        }

    def _upsert_por_id_linea_ruta(self, payload):
        self.cursor.execute(
            "SELECT id FROM config.rutas WHERE id_linea=%s AND UPPER(ruta_comercial)=%s",
            (payload["id_linea"], self._u(payload["ruta_comercial"])),
        )
        row = self.cursor.fetchone()
        if row:
            return "actualizado", self.actualizar_ruta(id=row["id"], **payload)
        else:
            return "insertado", self.crear_ruta(**payload)

    # ── Carga masiva efectiva ─────────────────────────────────────────────────
    def carga_masiva_csv_bytes(self, contenido: bytes):
        texto = contenido.decode("utf-8", errors="ignore")
        lector = csv.DictReader(io.StringIO(texto))
        ins, upd, errs = 0, 0, []
        for i, fila in enumerate(lector, start=2):
            try:
                payload = self._normalizar_payload_archivo(fila)
                accion, _ = self._upsert_por_id_linea_ruta(payload)
                if accion == "insertado":
                    ins += 1
                else:
                    upd += 1
            except Exception as e:
                errs.append({"row": i, "error": str(e)})
        return {"insertados": ins, "actualizados": upd, "errores": errs}

    def carga_masiva_xlsx_bytes(self, contenido: bytes):
        if not openpyxl:
            raise ValueError("openpyxl no está disponible en el entorno")
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        headers = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        ins, upd, errs = 0, 0, []
        for idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = {
                    headers[i]: (fila[i] if i < len(fila) else None)
                    for i in range(len(headers))
                }
                payload = self._normalizar_payload_archivo(data)
                accion, _ = self._upsert_por_id_linea_ruta(payload)
                if accion == "insertado":
                    ins += 1
                else:
                    upd += 1
            except Exception as e:
                errs.append({"row": idx, "error": str(e)})
        return {"insertados": ins, "actualizados": upd, "errores": errs}

    # ── Previsualización (sin escritura en DB) ────────────────────────────────
    def _preview_rows(self, rows):
        muestra, ins, upd, errs = [], 0, 0, []
        for i, data in rows:
            try:
                payload = self._normalizar_payload_archivo(data)
                self.cursor.execute(
                    "SELECT 1 FROM config.rutas WHERE id_linea=%s AND UPPER(ruta_comercial)=%s",
                    (payload["id_linea"], self._u(payload["ruta_comercial"])),
                )
                existe = self.cursor.fetchone() is not None
                accion = "actualizado" if existe else "insertado"
                if accion == "insertado":
                    ins += 1
                else:
                    upd += 1
                if len(muestra) < 50:
                    muestra.append(
                        {
                            "fila": i,
                            "id_linea": payload["id_linea"],
                            "ruta_comercial": payload["ruta_comercial"],
                            "cop_id": payload["id_cop"],
                            "accion": accion,
                        }
                    )
            except Exception as e:
                errs.append({"row": i, "error": str(e)})
        return {"insertados": ins, "actualizados": upd, "errores": errs, "muestra": muestra}

    def previsualizar_csv_bytes(self, contenido: bytes):
        texto = contenido.decode("utf-8", errors="ignore")
        lector = csv.DictReader(io.StringIO(texto))
        rows = list((i, r) for i, r in enumerate(lector, start=2))
        return self._preview_rows(rows)

    def previsualizar_xlsx_bytes(self, contenido: bytes):
        if not openpyxl:
            raise ValueError("openpyxl no está disponible en el entorno")
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        headers = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        rows = []
        for idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {
                headers[i]: (fila[i] if i < len(fila) else None)
                for i in range(len(headers))
            }
            rows.append((idx, data))
        return self._preview_rows(rows)

# ═══════════════════════════════════════════════════════════════════════════════
# GestionRutasKML — Trazados y paraderos KML para config.rutas_kml
# Sin dependencia de PostGIS — usa PostgreSQL nativo (TEXT + DOUBLE PRECISION)
# GeoJSON construido en Python y entregado directamente a Leaflet
# ═══════════════════════════════════════════════════════════════════════════════

class GestionRutasKML:
    """
    Gestión de trazados y paraderos KML por ruta (config.rutas_kml).
    No requiere PostGIS ni ninguna extensión — solo PostgreSQL estándar.
    Esquema unificado: todas las geometrías y metadatos van en coordenadas_json.
        Trazado:  {"type":"LineString","coordinates":[[lon,lat],...],"variante":N}
        Paradero: {"type":"Point","coordinates":[lon,lat],"secuencia":N,
                    "codigo":"XXX","nombre":"Nombre parada","variante":N}
    """

    def __init__(self):
        self.connection = get_db_pool().getconn()
        if not self.connection.closed:
            self.connection.rollback()
        self.connection.cursor_factory = RealDictCursor
        self.cursor = self.connection.cursor()
        # Crear tabla automáticamente si no existe — sin PostGIS
        self._asegurar_tabla()

    def cerrar_conexion(self):
        if getattr(self, "cursor", None):
            try:
                self.cursor.close()
            except Exception:
                pass
            self.cursor = None
        if getattr(self, "connection", None):
            try:
                if not self.connection.closed:
                    self.connection.rollback()
                get_db_pool().putconn(self.connection)
            except Exception:
                pass
            self.connection = None

    def _col_existe_kml(self, col: str) -> bool:
        self.cursor.execute(
            """
            SELECT 1 FROM information_schema.columns
            WHERE table_schema='config' AND table_name='rutas_kml'
                AND column_name=%s LIMIT 1
            """,
            (col,),
        )
        return self.cursor.fetchone() is not None

    def _asegurar_tabla(self):
        """
        Crea config.rutas_kml si no existe.

        Migración desde esquema anterior (columnas lat/lon/secuencia/... separadas):
        1. Reconstruye coordenadas_json para paraderos que aún lo tengan NULL
            usando las columnas lat, lon, secuencia, codigo_parada, nombre_parada,
            num_variante mientras todavía existen.
        2. Rellena coordenadas_json de trazados que faltasen (raro, pero seguro).
        3. Aplica NOT NULL en coordenadas_json.
        4. Elimina las columnas obsoletas.
        Sin PostGIS — solo PostgreSQL estándar.
        """
        self.cursor.execute(_DDL_RUTAS_KML)
        self.connection.commit()

        # ── Paso 1: migrar datos de paraderos si las columnas antiguas existen ──
        tiene_lat = self._col_existe_kml("lat")
        tiene_seq = self._col_existe_kml("secuencia")

        if tiene_lat:
            # Reconstruir coordenadas_json para paraderos con lat/lon válidos
            # que todavía no tengan JSON (NULL o '{}')
            self.cursor.execute(
                """
                UPDATE config.rutas_kml
                SET coordenadas_json = json_build_object(
                    'type',      'Point',
                    'coordinates', json_build_array(lon, lat),
                    'secuencia', COALESCE(secuencia, 0),
                    'codigo',    COALESCE(codigo_parada, ''),
                    'nombre',    COALESCE(nombre_parada, nombre_placemark, ''),
                    'variante',  COALESCE(num_variante, 1)
                )::text
                WHERE tipo = 'paradero'
                  AND lat  IS NOT NULL
                  AND lon  IS NOT NULL
                  AND (coordenadas_json IS NULL OR coordenadas_json = '{}');
                """
            )
        # ── Paso 2: paraderos sin lat (datos corruptos) → JSON mínimo vacío ────
        self.cursor.execute(
            """
            UPDATE config.rutas_kml
            SET coordenadas_json = '{"type":"Point","coordinates":[]}'
            WHERE tipo = 'paradero'
              AND (coordenadas_json IS NULL OR coordenadas_json = '{}');
            """
        )
        # ── Paso 3: trazados sin JSON (no debería ocurrir, pero por seguridad) ─
        self.cursor.execute(
            """
            UPDATE config.rutas_kml
            SET coordenadas_json = '{"type":"LineString","coordinates":[]}'
            WHERE tipo = 'trazado'
              AND (coordenadas_json IS NULL OR coordenadas_json = '{}');
            """
        )

        # ── Paso 4: aplicar NOT NULL ─────────────────────────────────────────
        self.cursor.execute(
            """
            SELECT is_nullable FROM information_schema.columns
            WHERE table_schema='config' AND table_name='rutas_kml'
              AND column_name='coordenadas_json' LIMIT 1
            """
        )
        row = self.cursor.fetchone()
        if row and row["is_nullable"] == "YES":
            self.cursor.execute(
                "ALTER TABLE config.rutas_kml "
                "ALTER COLUMN coordenadas_json SET NOT NULL;"
            )

        # ── Paso 5: eliminar columnas obsoletas ──────────────────────────────
        for col in ("num_variante", "secuencia", "codigo_parada",
                    "nombre_parada", "lat", "lon"):
            if self._col_existe_kml(col):
                self.cursor.execute(
                    f"ALTER TABLE config.rutas_kml DROP COLUMN IF EXISTS {col};"
                )

        self.connection.commit()

    # ── Carga de KML ──────────────────────────────────────────────────────────
    def cargar_kml(
        self,
        id_linea: int,
        contenido_bytes: bytes,
        nombre_archivo: str = "",
        usuario: str = "sistema",
        reemplazar: bool = True,
    ) -> dict:
        """
        Parsea el .kml y persiste trazados y paraderos en config.rutas_kml.
        Si reemplazar=True elimina los registros previos del id_linea.
        Retorna: { id_linea, trazados, paraderos, total, archivo }
        """
        registros = parsear_kml_bytes(contenido_bytes)
        if not registros:
            raise ValueError(
                "El KML no contiene placemarks válidos (LineString / Point)."
            )

        _INSERT = """
            INSERT INTO config.rutas_kml
                (id_linea, nombre_placemark, tipo,
                 coordenadas_json, total_puntos, archivo_origen, cargado_por)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        try:
            if reemplazar:
                self.cursor.execute(
                    "DELETE FROM config.rutas_kml WHERE id_linea = %s;",
                    (id_linea,),
                )
            cnt_t = cnt_p = 0
            for r in registros:
                self.cursor.execute(_INSERT, (
                    id_linea,
                    r["nombre_placemark"],
                    r["tipo"],
                    r["coordenadas_json"],
                    r["total_puntos"],
                    nombre_archivo,
                    usuario,
                ))
                if r["tipo"] == "trazado":
                    cnt_t += 1
                else:
                    cnt_p += 1
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise e

        return {
            "id_linea": id_linea,
            "trazados": cnt_t,
            "paraderos": cnt_p,
            "total": cnt_t + cnt_p,
            "archivo": nombre_archivo,
        }

    # ── Estado / resumen ──────────────────────────────────────────────────────
    def estado_kml(self, id_linea: int) -> dict:
        """Resumen de lo cargado: tiene_kml, conteos por tipo, fecha, archivo."""
        self.cursor.execute(
            """
            SELECT tipo,
                    COUNT(*)             AS cantidad,
                    MAX(created_at)      AS ultima_carga,
                    MAX(archivo_origen)  AS archivo,
                    MAX(cargado_por)     AS usuario
            FROM config.rutas_kml
            WHERE id_linea = %s
            GROUP BY tipo
            """,
            (id_linea,),
        )
        rows = self.cursor.fetchall()

        if not rows:
            return {"id_linea": id_linea, "tiene_kml": False}

        detalle = {}
        for r in rows:
            detalle[r["tipo"]] = {
                "cantidad": r["cantidad"],
                "ultima_carga": str(r["ultima_carga"]) if r["ultima_carga"] else None,
                "archivo": r["archivo"],
                "usuario": r["usuario"],
            }
        return {"id_linea": id_linea, "tiene_kml": True, "detalle": detalle}

    # ── GeoJSON para el mapa ──────────────────────────────────────────────────
    def trazados_geojson(self, id_linea: int) -> dict:
        """
        GeoJSON FeatureCollection con los trazados de la ruta.
        La variante y coordenadas se leen directamente de coordenadas_json.
        Listo para consumir con L.geoJSON() en Leaflet.
        """
        self.cursor.execute(
            """
            SELECT id, nombre_placemark, total_puntos, coordenadas_json
            FROM config.rutas_kml
            WHERE id_linea = %s AND tipo = 'trazado'
            ORDER BY id
            """,
            (id_linea,),
        )
        rows = self.cursor.fetchall()

        features = []
        for r in rows:
            try:
                geom = json.loads(r["coordenadas_json"])
            except (TypeError, json.JSONDecodeError):
                continue
            variante = geom.pop("variante", 1)  # extraer del JSON, no exponer en geometry
            features.append({
                "type": "Feature",
                "geometry": {"type": geom["type"], "coordinates": geom["coordinates"]},
                "properties": {
                    "id": r["id"],
                    "nombre": r["nombre_placemark"],
                    "nombre_placemark": r["nombre_placemark"],
                    "variante": variante,
                    "total_puntos": r["total_puntos"],
                },
            })
        return {"type": "FeatureCollection", "features": features}

    def paraderos_geojson(self, id_linea: int) -> dict:
        """
        GeoJSON FeatureCollection con los paraderos de la ruta.
        Coordenadas y metadatos (secuencia, codigo, nombre, variante)
        se leen de coordenadas_json — no hay columnas lat/lon separadas.
        """
        self.cursor.execute(
            """
            SELECT id, nombre_placemark, coordenadas_json
            FROM config.rutas_kml
            WHERE id_linea = %s AND tipo = 'paradero'
            ORDER BY id
            """,
            (id_linea,),
        )
        rows = self.cursor.fetchall()

        features = []
        for r in rows:
            try:
                geom = json.loads(r["coordenadas_json"])
            except (TypeError, json.JSONDecodeError):
                continue
            coords = geom.get("coordinates")
            if not coords or len(coords) < 2:
                continue
            features.append({
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": coords},
                "properties": {
                    "id": r["id"],
                    "nombre_placemark": r["nombre_placemark"],
                    "variante": geom.get("variante", 1),
                    "secuencia": geom.get("secuencia"),
                    "codigo": geom.get("codigo"),
                    "nombre": geom.get("nombre"),
                },
            })
        return {"type": "FeatureCollection", "features": features}

    # ── Eliminar ──────────────────────────────────────────────────────────────
    def eliminar_kml(self, id_linea: int) -> int:
        """Elimina todos los registros KML del id_linea. Retorna cantidad eliminada."""
        try:
            self.cursor.execute(
                "DELETE FROM config.rutas_kml WHERE id_linea = %s;",
                (id_linea,),
            )
            eliminados = self.cursor.rowcount
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise e
        return eliminados

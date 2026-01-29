import os, io, csv
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2 import errors
from datetime import datetime
from dotenv import load_dotenv
from zoneinfo import ZoneInfo

try:
    import openpyxl  # Para leer .xlsx
except Exception:
    openpyxl = None

# Variables de entorno
load_dotenv()
DATABASE_PATH = os.getenv("DATABASE_PATH")
TZ_BOGOTA = ZoneInfo("America/Bogota")

def ahora_bogota() -> datetime:
    return datetime.now(TZ_BOGOTA)

class GestionBuses:
    def __init__(self):
        self.connection = psycopg2.connect(DATABASE_PATH, options='-c timezone=America/Bogota')
        self.cursor = self.connection.cursor(cursor_factory=RealDictCursor)
        with self.connection.cursor() as c:
            c.execute("SET TIME ZONE 'America/Bogota';")
        self.connection.commit()

    def cerrar_conexion(self):
        if getattr(self, "cursor", None):
            self.cursor.close()
        if getattr(self, "connection", None):
            self.connection.close()

    # === Helpers para detectar columnas y armar joins dinámicos ===
    def _col_exists(self, schema: str, table: str, column: str) -> bool:
        self.cursor.execute("""
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema=%s AND table_name=%s AND column_name=%s
            LIMIT 1
        """, (schema, table, column))
        return self.cursor.fetchone() is not None

    def _cop_joins_and_exprs(self):
        """
        Devuelve (joins_sql, comp_expr, zona_expr) adaptado a tu estructura actual:
        - config.cop.id_componente → config.componente.componente
        - config.cop.id_zona → config.zona.zona
        """
        has_id_comp = self._col_exists('config', 'cop', 'id_componente')
        has_id_zona = self._col_exists('config', 'cop', 'id_zona')

        joins = []
        comp_expr = "NULL"
        zona_expr = "NULL"

        if has_id_comp:
            joins.append("LEFT JOIN config.componente comp ON comp.id = c.id_componente")
            comp_expr = "comp.componente"

        if has_id_zona:
            joins.append("LEFT JOIN config.zona z ON z.id = c.id_zona")
            zona_expr = "z.zona"

        return (" ".join(joins), comp_expr, zona_expr)

    # ---------- Utilidades ----------
    def _fila_o_error(self, cur):
        fila = cur.fetchone()
        if not fila:
            raise ValueError("Registro no encontrado")
        return fila

    def _paginacion(self, pagina:int, tamano:int):
        off = (pagina - 1) * tamano
        return " LIMIT %s OFFSET %s ", [tamano, off]

    def _filtro_busqueda(self, q: str):
        if not q:
            return "", []
        qn = f"%{q.strip().upper()}%"
        return " AND (UPPER(b.placa) LIKE %s OR UPPER(COALESCE(b.no_interno,'')) LIKE %s) ", [qn, qn]

    def _u(self, s):
        if s is None:
            return None
        return str(s).strip().upper()

    def _resolver_id_cop(self, id_cop=None, nombre_cop=None):
        """
        Modo híbrido:
        - Si llega id_cop (numérico), lo usa tal cual.
        - Si llega nombre_cop, busca en config.cop.cop por nombre exacto (case-insensitive).
        """
        if id_cop:
            return int(id_cop)
        if nombre_cop:
            self.cursor.execute(
                "SELECT id FROM config.cop WHERE UPPER(cop) = %s",
                (nombre_cop.strip().upper(),)
            )
            row = self.cursor.fetchone()
            if not row:
                raise ValueError(f"COP '{nombre_cop}' no existe")
            return int(row["id"])
        raise ValueError("Centro de Operación no especificado (ID COP o nombre)")

    def _normalizar_combustible(self, c):
        """
        Normaliza combustible a {ACPM, GNV, ELECTRICO}.
        Acepta 'ELÉCTRICO' (con tilde) y lo convierte a 'ELECTRICO'.
        """
        cu = self._u(c) or ""
        if cu in ("ACPM", "GNV", "ELECTRICO"):
            return cu
        if cu == "ELÉCTRICO":
            return "ELECTRICO"
        raise ValueError("Combustible inválido (permitidos: ACPM, GNV, ELECTRICO)")

    # ---------- Listar buses ----------
    def listar_buses(self, q=None, estado=None, id_cop=None, combustible=None,
                    tipologia=None, modelo_desde=None, modelo_hasta=None,
                    pagina:int=1, tamano:int=10):
        where, params = " WHERE 1=1 ", []
        s, sp = self._filtro_busqueda(q)
        where += s; params += sp

        if estado is not None:
            where += " AND b.estado=%s "; params.append(int(estado))
        if id_cop:
            where += " AND b.id_cop=%s "; params.append(int(id_cop))
        if combustible:
            where += " AND b.combustible=%s "; params.append(self._u(combustible))
        if tipologia:
            where += " AND UPPER(b.tipologia)=%s "; params.append(self._u(tipologia))
        if modelo_desde:
            where += " AND b.modelo >= %s "; params.append(int(modelo_desde))
        if modelo_hasta:
            where += " AND b.modelo <= %s "; params.append(int(modelo_hasta))

        joins_cop, comp_expr, zona_expr = self._cop_joins_and_exprs()

        sql_count = f"""
            SELECT COUNT(1)
            FROM config.buses_cexp b
            JOIN config.cop c ON c.id = b.id_cop
            {where}
        """
        self.cursor.execute(sql_count, params)
        total = self.cursor.fetchone()["count"]

        order = " ORDER BY b.id DESC "
        pag, pp = self._paginacion(pagina, tamano)
        sqlq = f"""
            SELECT b.id, b.placa, b.no_interno, b.tipologia, b.modelo, b.marca, b.linea, b.carroceria,
                b.combustible, b.tecnologia, b.estado, b.id_cop,
                c.cop,
                {comp_expr} AS componente,
                {zona_expr} AS zona,
                to_char(b.created_at,'YYYY-MM-DD HH24:MI') as created_at,
                to_char(b.updated_at,'YYYY-MM-DD HH24:MI') as updated_at
            FROM config.buses_cexp b
            JOIN config.cop c ON c.id = b.id_cop
            {joins_cop}
            {where} {order} {pag}
        """
        self.cursor.execute(sqlq, params + pp)
        return self.cursor.fetchall(), total

    # ---------- Crear / Actualizar / Estado ----------
    def crear_bus(self, placa:str, combustible:str, id_cop:int, estado:int=1,
                  no_interno:str=None, tipologia:str=None, modelo:int=None,
                  marca:str=None, linea:str=None, carroceria:str=None, tecnologia:str=None):
        placa_u = self._u(placa)
        if not placa_u:
            raise ValueError("La placa es obligatoria")

        combustible_n = self._normalizar_combustible(combustible)

        try:
            sqlq = """
                INSERT INTO config.buses_cexp
                (placa, no_interno, tipologia, modelo, marca, linea, carroceria, combustible, tecnologia, id_cop, estado)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id, placa, estado, id_cop,
                          to_char(created_at,'YYYY-MM-DD HH24:MI') as created_at,
                          to_char(updated_at,'YYYY-MM-DD HH24:MI') as updated_at
            """
            self.cursor.execute(sqlq, (
                placa_u,
                self._u(no_interno) if no_interno else None,
                self._u(tipologia) if tipologia else None,
                int(modelo) if modelo else None,
                self._u(marca) if marca else None,
                self._u(linea) if linea else None,
                self._u(carroceria) if carroceria else None,
                combustible_n,
                self._u(tecnologia) if tecnologia else None,
                int(id_cop),
                int(estado)
            ))
            fila = self.cursor.fetchone()
            self.connection.commit()
            return fila
        except errors.ForeignKeyViolation:
            self.connection.rollback(); raise ValueError("El Centro de Operación (COP) no existe")
        except errors.UniqueViolation as e:
            self.connection.rollback()
            s = str(e).lower()
            if "placa" in s: raise ValueError("La placa ya existe")
            if "no_interno" in s: raise ValueError("El No Interno ya existe")
            raise
        except Exception as e:
            self.connection.rollback(); raise e

    def actualizar_bus(self, id:int, placa:str, combustible:str, id_cop:int, estado:int=1,
                    no_interno:str=None, tipologia:str=None, modelo:int=None,
                    marca:str=None, linea:str=None, carroceria:str=None, tecnologia:str=None):
        placa_u = self._u(placa)
        if not placa_u:
            raise ValueError("La placa es obligatoria")

        combustible_n = self._normalizar_combustible(combustible)

        try:
            sqlq = """
                UPDATE config.buses_cexp
                SET placa=%s, no_interno=%s, tipologia=%s, modelo=%s, marca=%s, linea=%s, carroceria=%s,
                    combustible=%s, tecnologia=%s, id_cop=%s, estado=%s, updated_at=now()
                WHERE id=%s
                RETURNING id, placa, estado, id_cop,
                        to_char(created_at,'YYYY-MM-DD HH24:MI') as created_at,
                        to_char(updated_at,'YYYY-MM-DD HH24:MI') as updated_at
            """
            self.cursor.execute(sqlq, (
                placa_u,
                self._u(no_interno) if no_interno else None,
                self._u(tipologia) if tipologia else None,
                int(modelo) if modelo else None,
                self._u(marca) if marca else None,
                self._u(linea) if linea else None,
                self._u(carroceria) if carroceria else None,
                combustible_n,
                self._u(tecnologia) if tecnologia else None,
                int(id_cop),
                int(estado),
                int(id)
            ))
            fila = self._fila_o_error(self.cursor)
            self.connection.commit()
            return fila
        except errors.ForeignKeyViolation:
            self.connection.rollback(); raise ValueError("El Centro de Operación (COP) no existe")
        except errors.UniqueViolation as e:
            self.connection.rollback()
            s = str(e).lower()
            if "placa" in s: raise ValueError("La placa ya existe")
            if "no_interno" in s: raise ValueError("El No Interno ya existe")
            raise
        except Exception as e:
            self.connection.rollback(); raise e

    def cambiar_estado(self, id:int, estado:int):
        self.cursor.execute("""
            UPDATE config.buses_cexp
            SET estado=%s, updated_at=now()
            WHERE id=%s
            RETURNING id, placa, estado,
                      to_char(created_at,'YYYY-MM-DD HH24:MI') as created_at,
                      to_char(updated_at,'YYYY-MM-DD HH24:MI') as updated_at
        """, (int(estado), int(id)))
        fila = self._fila_o_error(self.cursor)
        self.connection.commit()
        return fila

    # ---------- Datos de apoyo (COP / filtros) ----------
    def listar_cop(self, componente: str = None, zona: str = None, estado: int = 1):
        joins_cop, comp_expr, zona_expr = self._cop_joins_and_exprs()
        where, params = " WHERE 1=1 ", []
        if estado is not None:
            where += " AND c.estado=%s "; params.append(int(estado))
        if componente:
            where += f" AND UPPER({comp_expr})=%s "; params.append(self._u(componente))
        if zona:
            where += f" AND UPPER({zona_expr})=%s "; params.append(self._u(zona))

        sql = f"""
            SELECT c.id, c.cop,
                {comp_expr} AS componente,
                {zona_expr} AS zona
            FROM config.cop c
            {joins_cop}
            {where}
            ORDER BY {comp_expr}, {zona_expr}, c.cop
        """
        self.cursor.execute(sql, params)
        return self.cursor.fetchall()

    def listar_componentes(self):
        joins_cop, comp_expr, _ = self._cop_joins_and_exprs()
        sql = f"""
            SELECT DISTINCT {comp_expr} AS componente
            FROM config.cop c
            {joins_cop}
            WHERE c.estado=1 AND {comp_expr} IS NOT NULL
            ORDER BY {comp_expr}
        """
        self.cursor.execute(sql)
        return [r["componente"] for r in self.cursor.fetchall() if r["componente"]]

    def listar_zonas(self, componente: str = None):
        joins_cop, comp_expr, zona_expr = self._cop_joins_and_exprs()
        where, params = " WHERE c.estado=1 ", []
        if componente:
            where += f" AND UPPER({comp_expr})=%s "; params.append(self._u(componente))
        sql = f"""
            SELECT DISTINCT {zona_expr} AS zona
            FROM config.cop c
            {joins_cop}
            {where} AND {zona_expr} IS NOT NULL
            ORDER BY {zona_expr}
        """
        self.cursor.execute(sql, params)
        return [r["zona"] for r in self.cursor.fetchall() if r["zona"]]

    # ---------- Normalización carga masiva ----------
    def _normalizar_payload_archivo(self, d: dict):
        """
        Cabeceras esperadas (flexibles):
        - Placa, No Interno, Tipologia, Modelo, Marca, Linea, Carroceria, Combustible, Tecnologia,
        - Centro Operacion  (nombre COP)   O   ID COP / id_cop (numérico)
        """
        def get(k):
            v = d.get(k)
            return v if v is None else str(v).strip()

        payload = {
            "placa": get("Placa"),
            "no_interno": get("No Interno") or None,
            "tipologia": get("Tipologia") or None,
            "modelo": int(get("Modelo")) if (get("Modelo") and str(get("Modelo")).isdigit()) else None,
            "marca": get("Marca") or None,
            "linea": get("Linea") or None,
            "carroceria": get("Carroceria") or None,
            "combustible": self._normalizar_combustible(get("Combustible")),
            "tecnologia": get("Tecnologia") or None,
            "id_cop": None,
            "estado": 1
        }

        # 1) Prioridad: ID COP si llega
        id_cop_val = get("ID COP") or get("id_cop") or get("Centro Operacion")
        if id_cop_val and id_cop_val.isdigit():
            payload["id_cop"] = int(id_cop_val)
        else:
            # 2) Resolver por nombre de COP
            nombre_cop = get("Centro Operacion")
            if nombre_cop:
                payload["id_cop"] = self._resolver_id_cop(nombre_cop=nombre_cop)

        if not payload["id_cop"]:
            raise ValueError("Centro Operacion/ID COP vacío o inválido")

        return payload

    def _upsert_por_placa(self, payload):
        self.cursor.execute("SELECT id FROM config.buses_cexp WHERE placa=%s", (self._u(payload["placa"]),))
        row = self.cursor.fetchone()
        if row:
            return "actualizado", self.actualizar_bus(id=row["id"], **payload)
        else:
            return "insertado", self.crear_bus(**payload)

    # ---------- Carga masiva (efectiva) ----------
    def carga_masiva_csv_bytes(self, contenido: bytes):
        texto = contenido.decode("utf-8", errors="ignore")
        lector = csv.DictReader(io.StringIO(texto))
        ins, upd, errs = 0, 0, []
        for i, fila in enumerate(lector, start=2):
            try:
                payload = self._normalizar_payload_archivo(fila)
                accion, _ = self._upsert_por_placa(payload)
                if accion == "insertado": ins += 1
                else: upd += 1
            except Exception as e:
                errs.append({"row": i, "error": str(e)})
        return {"insertados": ins, "actualizados": upd, "errores": errs}

    def carga_masiva_xlsx_bytes(self, contenido: bytes):
        if not openpyxl:
            raise ValueError("openpyxl no está disponible en el entorno")
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ins, upd, errs = 0, 0, []
        for idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = {headers[i]: (fila[i] if i < len(fila) else None) for i in range(len(headers))}
                # Excel puede entregar 1.0 en "Centro Operacion"; limpiar:
                if "Centro Operacion" in data and data["Centro Operacion"] is not None:
                    cv = str(data["Centro Operacion"]).strip()
                    if cv.endswith(".0"): cv = cv[:-2]
                    data["Centro Operacion"] = cv
                payload = self._normalizar_payload_archivo(data)
                accion, _ = self._upsert_por_placa(payload)
                if accion == "insertado": ins += 1
                else: upd += 1
            except Exception as e:
                errs.append({"row": idx, "error": str(e)})
        return {"insertados": ins, "actualizados": upd, "errores": errs}

    # ---------- Previsualización (no escribe en DB) ----------
    def _preview_rows(self, rows):
        muestra = []
        ins, upd, errs = 0, 0, []
        for i, data in rows:
            try:
                payload = self._normalizar_payload_archivo(data)
                # Chequear existencia por placa
                self.cursor.execute("SELECT 1 FROM config.buses_cexp WHERE placa=%s", (self._u(payload["placa"]),))
                existe = self.cursor.fetchone() is not None
                accion = "actualizado" if existe else "insertado"
                if accion == "insertado": ins += 1
                else: upd += 1
                if len(muestra) < 50:
                    muestra.append({
                        "fila": i,
                        "placa": self._u(payload["placa"]),
                        "cop_id": payload["id_cop"],
                        "accion": accion,
                        "combustible": payload["combustible"],
                        "tipologia": self._u(payload.get("tipologia")),
                        "modelo": payload.get("modelo")
                    })
            except Exception as e:
                errs.append({"row": i, "error": str(e)})
        return {"insertados": ins, "actualizados": upd, "errores": errs, "muestra": muestra}

    def previsualizar_csv_bytes(self, contenido: bytes):
        texto = contenido.decode("utf-8", errors="ignore")
        lector = csv.DictReader(io.StringIO(texto))
        rows = list((i, r) for i, r in enumerate(lector, start=2))
        return self._preview_rows(rows)

    def previsualizar_xlsx_bytes(self, contenido: bytes):
        if not openpyxl:
            raise ValueError("openpyxl no está disponible en el entorno")
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {headers[i]: (fila[i] if i < len(fila) else None) for i in range(len(headers))}
            if "Centro Operacion" in data and data["Centro Operacion"] is not None:
                cv = str(data["Centro Operacion"]).strip()
                if cv.endswith(".0"): cv = cv[:-2]
                data["Centro Operacion"] = cv
            rows.append((idx, data))
        return self._preview_rows(rows)

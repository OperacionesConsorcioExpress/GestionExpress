from azure.storage.blob import BlobServiceClient
from model.gestion_usuarios import HandleDB, Cargue_Roles_Blob_Storage
from datetime import datetime, timedelta
import os

# ─────────────────────────────────────────────────────────────────
# Caché en memoria a nivel de clase (compartido entre instancias)
# TTL configurable — por defecto 5 minutos
# ─────────────────────────────────────────────────────────────────
_CACHE: dict = {}
CACHE_TTL_MINUTES = 5

class ContainerModel:
    def __init__(self):
        connection_string = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        self.blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        self.db = HandleDB()

    # ─────────────────────────────────────────────
    # Helpers de caché
    # ─────────────────────────────────────────────
    def _cache_get(self, key: str):
        """Retorna datos del caché si existen y no han expirado, si no None."""
        entry = _CACHE.get(key)
        if entry and datetime.now() < entry["expires"]:
            return entry["data"]
        return None

    def _cache_set(self, key: str, data):
        """Guarda datos en el caché con TTL."""
        _CACHE[key] = {
            "data":    data,
            "expires": datetime.now() + timedelta(minutes=CACHE_TTL_MINUTES)
        }

    def _cache_invalidate(self, container_name: str):
        """Invalida el caché de un contenedor (subida, eliminación, refresco)."""
        _CACHE.pop(container_name, None)

    # ─────────────────────────────────────────────
    # Contenedores
    # ─────────────────────────────────────────────
    def get_containers(self) -> list:
        """Lista todos los contenedores disponibles en Azure Blob Storage."""
        return [c.name for c in self.blob_service_client.list_containers()]

    def create_container(self, name: str):
        """Crea un nuevo contenedor."""
        self.blob_service_client.create_container(name)

    def get_allowed_containers(self, user_rol_storage: str) -> list:
        """Filtra los contenedores accesibles para el rol del usuario."""
        roles     = Cargue_Roles_Blob_Storage()
        asignados = roles.get_contenedores_por_rol(user_rol_storage)
        return [c for c in self.get_containers() if c in asignados]

    def get_acciones_permitidas(self, user_rol_storage) -> dict:
        """Retorna dict con booleanos de acciones permitidas para el rol del usuario."""
        roles = Cargue_Roles_Blob_Storage()
        return roles.get_acciones_por_rol(user_rol_storage)

    # ─────────────────────────────────────────────
    # Árbol de archivos — endpoint unificado
    # ─────────────────────────────────────────────
    def get_files(self, container_name: str) -> dict:
        """
        Retorna solo el árbol jerárquico (compatibilidad con /files).
        Internamente reutiliza get_files_with_metadata para aprovechar el caché.
        """
        result = self.get_files_with_metadata(container_name)
        return result["tree"]

    def get_files_with_metadata(self, container_name: str) -> dict:
        """
        Retorna en una sola pasada:
            - tree:     árbol jerárquico { carpeta: { archivo: None } }
            - metadata: { "ruta/archivo": { size, last_modified, content_type, etag } }
            - stats:    { total_files, total_size, total_folders }

        Optimizaciones aplicadas:
          1. Caché en memoria (TTL 5 min) — segunda consulta es instantánea
          2. results_per_page=5000 — mínimo de roundtrips de paginación a Azure
          3. Un único recorrido de la lista para construir árbol + metadata + stats
          4. get_files() reutiliza este método — Azure solo se consulta una vez
             aunque el frontend llame a /files y /files-metadata en paralelo
        """
        # ── 1. Retornar desde caché si está vigente ──
        cached = self._cache_get(container_name)
        if cached is not None:
            return cached

        # ── 2. Consultar Azure con paginación máxima ──
        container_client = self.blob_service_client.get_container_client(container_name)
        blobs = list(container_client.list_blobs(results_per_page=5000))

        # ── 3. Un solo recorrido: árbol + metadata + stats ──
        tree     = {}
        metadata = {}
        folders  = set()
        total_sz = 0

        for blob in blobs:
            parts = blob.name.split('/')
            level = tree

            for i, part in enumerate(parts[:-1]):
                folders.add('/'.join(parts[:i + 1]))
                if part not in level or level[part] is None:
                    level[part] = {}
                level = level[part]

            level[parts[-1]] = None

            size      = blob.size or 0
            total_sz += size

            ct = None
            try:
                ct = blob.content_settings.content_type if blob.content_settings else None
            except Exception:
                pass

            metadata[blob.name] = {
                "size":          size,
                "last_modified": blob.last_modified.isoformat() if blob.last_modified else None,
                "content_type":  ct,
                "etag":          blob.etag,
                "blob_type":     str(blob.blob_type) if blob.blob_type else "BlockBlob",
            }

        result = {
            "tree":     tree,
            "metadata": metadata,
            "stats": {
                "total_files":   len(blobs),
                "total_size":    total_sz,
                "total_folders": len(folders),
            }
        }

        # ── 4. Guardar en caché ──
        self._cache_set(container_name, result)
        return result

    # ─────────────────────────────────────────────
    # Descarga
    # ─────────────────────────────────────────────
    def download_file(self, container_name: str, file_name: str) -> bytes:
        """Descarga el contenido binario completo (uso interno o archivos pequeños)."""
        blob_client = self.blob_service_client.get_blob_client(
            container=container_name, blob=file_name
        )
        return blob_client.download_blob().readall()

    def stream_file(self, container_name: str, file_name: str):
        """
        Generador: streaming directo Azure -> cliente en chunks.
        max_concurrency=4 abre 4 conexiones paralelas a Azure para maximizar throughput.
        Nunca carga el archivo completo en RAM.
        """
        blob_client = self.blob_service_client.get_blob_client(
            container=container_name, blob=file_name
        )
        stream = blob_client.download_blob(max_concurrency=4)
        for chunk in stream.chunks():
            yield chunk

    def get_blob_size(self, container_name: str, file_name: str) -> int:
        """Retorna el tamaño del blob en bytes sin descargarlo."""
        # Si el blob está en el caché de metadata, evitar llamada a Azure
        cached = self._cache_get(container_name)
        if cached and file_name in cached.get("metadata", {}):
            return cached["metadata"][file_name].get("size", 0)
        blob_client = self.blob_service_client.get_blob_client(
            container=container_name, blob=file_name
        )
        return blob_client.get_blob_properties().size

    # ─────────────────────────────────────────────
    # Operaciones que invalidan el caché
    # ─────────────────────────────────────────────
    def delete_file(self, container_name: str, file_name: str):
        """Elimina un blob e invalida el caché del contenedor."""
        blob_client = self.blob_service_client.get_blob_client(
            container=container_name, blob=file_name
        )
        blob_client.delete_blob()
        self._cache_invalidate(container_name)

    async def upload_file(self, container_name: str, file_name: str, file_content: bytes):
        """Sube un archivo e invalida el caché del contenedor."""
        blob_client = self.blob_service_client.get_blob_client(
            container=container_name, blob=file_name
        )
        await blob_client.upload_blob(file_content, overwrite=True)
        self._cache_invalidate(container_name)

    # ─────────────────────────────────────────────
    # Propiedades detalladas de un blob
    # ─────────────────────────────────────────────
    def get_blob_properties(self, container_name: str, blob_name: str) -> dict:
        """Propiedades detalladas de un blob (modal de propiedades en el frontend)."""
        blob_client = self.blob_service_client.get_blob_client(
            container=container_name, blob=blob_name
        )
        p = blob_client.get_blob_properties()
        return {
            "name":             blob_name,
            "container":        container_name,
            "size":             p.size,
            "last_modified":    p.last_modified.isoformat()  if p.last_modified  else None,
            "created_on":       p.creation_time.isoformat()  if p.creation_time  else None,
            "content_type":     p.content_settings.content_type     if p.content_settings else None,
            "content_encoding": p.content_settings.content_encoding if p.content_settings else None,
            "etag":             p.etag,
            "blob_type":        str(p.blob_type) if p.blob_type else None,
            "lease_status":     str(p.lease.status) if p.lease else None,
            "custom_metadata":  p.metadata or {},
        }

    # ─────────────────────────────────────────────
    # Preview paginado de Excel (visor sin descarga)
    # ─────────────────────────────────────────────
    def preview_excel(self, container_name: str, blob_name: str,
                        sheet: int = 0, page: int = 1, limit: int = 300) -> dict:
        """
        Retorna JSON paginado de una hoja Excel usando openpyxl (read_only=True).
        El browser NUNCA descarga el binario completo para previsualizar.
        Solo soporta .xlsx / .xlsm (openpyxl no procesa el formato binario .xls).
        """
        from openpyxl import load_workbook
        from io import BytesIO

        ext = blob_name.rsplit('.', 1)[-1].lower() if '.' in blob_name else ''
        if ext not in ('xlsx', 'xlsm', 'xlam'):
            return {
                "error":      f"Vista previa no disponible para archivos .{ext}. Descarga el archivo para abrirlo en Excel.",
                "sheets":     [],
                "rows":       [],
                "total_rows": 0,
                "total_cols": 0,
                "page":       page,
                "limit":      limit,
                "has_more":   False,
            }

        blob_client = self.blob_service_client.get_blob_client(
            container=container_name, blob=blob_name
        )
        data = blob_client.download_blob().readall()

        wb          = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        if sheet >= len(sheet_names):
            sheet = 0

        ws          = wb.worksheets[sheet]
        total_rows  = ws.max_row    or 0
        total_cols  = ws.max_column or 0

        offset    = (page - 1) * limit
        rows_data = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < offset:
                continue
            if i >= offset + limit:
                break
            rows_data.append(['' if v is None else str(v) for v in row])

        wb.close()

        return {
            "sheets":     list(sheet_names),
            "sheet":      sheet,
            "page":       page,
            "limit":      limit,
            "total_rows": total_rows,
            "total_cols": total_cols,
            "rows":       rows_data,
            "has_more":   (offset + limit) < total_rows,
        }